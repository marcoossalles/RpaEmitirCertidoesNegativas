from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import logging
import os

# Configura logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Define locale para português (ajusta os nomes dos meses)
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')  # Linux/macOS
except:
    locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')  # Windows

class PlanilhaMensalDuplicador:
    def __init__(self, caminho_arquivo=os.getenv('PASTA_PLANILHA_CNPJS')):
        """
        Inicializa a classe com o caminho do arquivo Excel e carrega a planilha.
        """
        try:
            logging.info(f"Carregando planilha: {caminho_arquivo}")
            self.caminho_arquivo = caminho_arquivo
            self.wb = load_workbook(caminho_arquivo)
            logging.info(f"Planilha carregada: {caminho_arquivo}")
        except Exception as e:
            logging.error(f"Erro ao carregar a planilha: {e}")
            exit()

    def duplicar_aba_mensal(self, 
                            colunas_para_limpar=("EMPRESAS", "CNPJ", "Simples Nacional", "Grupo", "Status", 
                                                "Inscrição", "Certidão", "FGTS", "TRABALHISTA", "Status Procesamento"),
                            linha_titulo=8, 
                            linha_dados=9):
        """
        Duplica a aba do mês atual para o próximo mês, limpando colunas específicas.
        """
        #locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

        hoje = datetime.today()

        # Nome da aba referente ao mês atual
        mes_atual_nome = hoje.strftime('%B').capitalize()
        ano_atual = hoje.year
        aba_mes_atual = f"{mes_atual_nome} {ano_atual}"

        # Calcular mês anterior
        mes_anterior = (hoje.replace(day=1) - timedelta(days=1))
        mes_anterior_nome = mes_anterior.strftime('%B').capitalize()
        ano_anterior = mes_anterior.year
        aba_mes_anterior = f"{mes_anterior_nome} {ano_anterior}"

        # Verifica se já existe a aba do mês atual
        if aba_mes_atual in self.wb.sheetnames:
            logging.info(f"Aba '{aba_mes_atual}' já existe. Nenhuma duplicação foi feita.")
            return

        # Verifica se a aba do mês anterior existe
        if aba_mes_anterior not in self.wb.sheetnames:
            logging.error(f"Aba de origem '{aba_mes_anterior}' não encontrada.")
            raise ValueError(f"Aba de origem '{aba_mes_anterior}' não encontrada.")

        logging.info(f"Duplicando aba '{aba_mes_anterior}' para '{aba_mes_atual}'.")

        aba_original = self.wb[aba_mes_anterior]
        nova_aba = self.wb.copy_worksheet(aba_original)
        nova_aba.title = aba_mes_atual

        # Prepara as colunas que devem ser limpas
        colunas_para_limpar = [col.strip().upper() for col in colunas_para_limpar]

        # Encontra índices das colunas a limpar
        titulos = [str(cell.value).strip().upper() for cell in aba_original[linha_titulo]]
        indices_para_limpar = [i + 1 for i, titulo in enumerate(titulos) if titulo in colunas_para_limpar]

        # Limpa os valores das colunas especificadas, a partir da linha de dados
        for row in nova_aba.iter_rows(min_row=linha_dados, max_row=nova_aba.max_row):
            for cell in row:
                if cell.column in indices_para_limpar:
                    cell.value = None

        self.wb.save(self.caminho_arquivo)
        logging.info(f"Aba criada com sucesso com base na aba '{aba_mes_atual}'.")

    def ler_aba_mes_atual(self, linha_titulo=8, linha_dados=9):
        """
        Lê os dados da aba do mês atual e retorna uma lista de dicionários com os dados.
        """
        hoje = datetime.today()
        mes_nome = hoje.strftime('%B').capitalize()
        ano = hoje.year
        nome_aba = f"{mes_nome} {ano}"

        if nome_aba not in self.wb.sheetnames:
            logging.error(f"Aba '{nome_aba}' não encontrada na planilha.")
            raise ValueError(f"Aba '{nome_aba}' não encontrada na planilha.")

        logging.info(f"Lendo dados da aba '{nome_aba}'.")

        aba = self.wb[nome_aba]

        # Coleta os nomes das colunas
        colunas = [cell.value for cell in aba[linha_titulo]]

        dados = []
        # Lê os dados a partir da linha de dados
        for row in aba.iter_rows(min_row=linha_dados, max_row=aba.max_row, values_only=True):
            if any(row):  # ignora linhas totalmente vazias
                linha_dict = {colunas[i]: row[i] for i in range(len(colunas)) if i < len(row)}
                dados.append(linha_dict)

        logging.info(f"{len(dados)} linhas de dados carregadas da aba '{nome_aba}'.")
        return dados
    
    def escrever_status_linha(self, linha, dicionario_status, linha_titulo=8):
        """
        Escreve os valores do dicionário de status nas colunas correspondentes da linha informada.
        """
        hoje = datetime.today()
        mes_nome = hoje.strftime('%B').capitalize()
        ano = hoje.year
        nome_aba = f"{mes_nome} {ano}"

        if nome_aba not in self.wb.sheetnames:
            logging.error(f"Aba '{nome_aba}' não encontrada no arquivo.")
            raise ValueError(f"Aba '{nome_aba}' não encontrada no arquivo.")

        logging.info(f"Escrevendo status na linha {linha} da aba '{nome_aba}'.")

        aba = self.wb[nome_aba]

        # Cria mapeamento entre nome da coluna e índice
        mapa_colunas = {}
        for cell in aba[linha_titulo]:
            if cell.value:
                nome_coluna = str(cell.value).strip()
                mapa_colunas[nome_coluna.lower()] = cell.column

        # Escreve os valores nas colunas correspondentes
        for nome_coluna, valor in dicionario_status.items():
            col = mapa_colunas.get(nome_coluna.lower())
            if col is None:
                logging.error(f"Coluna '{nome_coluna}' não encontrada na aba '{nome_aba}'.")
                raise ValueError(f"Coluna '{nome_coluna}' não encontrada na aba '{nome_aba}'.")
            aba.cell(row=linha, column=col, value=valor)

        self.wb.save(self.caminho_arquivo)
        logging.info(f"Status atualizado na linha {linha} da aba '{nome_aba}'.")
